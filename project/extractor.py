"""
extractor.py — извлечение содержимого файлов по типу.

Правила:
- .txt / .md / .csv / .html        → читать напрямую (без внешних инструментов)
- .pdf / .docx / .pptx             → Docling (текст)
- .pdf (при process_embedded_images: true)
                                   → PyMuPDF: изображения ≥ 100×100 px → Qwen3-VL
- .jpg / .png / и др.              → Qwen3-VL через Ollama (один проход: OCR + описание)
  Если process_standalone_images: false → вернуть пустой текст (только метаданные)
"""

import base64
import json
import re
from pathlib import Path

TEXT_EXTENSIONS        = {".txt", ".md", ".csv", ".html", ".htm"}
DOC_EXTENSIONS         = {".pdf", ".docx", ".pptx"}
SPREADSHEET_EXTENSIONS = {".xlsx", ".xls"}
IMAGE_EXTENSIONS       = {".jpg", ".jpeg", ".png", ".webp", ".tiff", ".tif"}

# Минимальный размер embedded-изображения для обработки (px по каждой стороне).
# Отсеивает иконки, разделители, декоративные элементы.
_MIN_IMAGE_PX = 100
_MAX_IMAGES_PER_PDF = 10
_MAX_PAGES_PER_PDF  = 10
_IMAGE_MAX_SIZE = 1280  # max px по длинной стороне перед отправкой в VLM


def extract(file_info: dict, config: dict) -> dict:
    """
    Возвращает единый текстовый объект:
    {
        path, ext, hash,
        raw_text,          # текст из документа (Docling / plain read)
        ocr_text,          # текст на изображении (Qwen3-VL, standalone)
        image_description, # описание сцены (Qwen3-VL, standalone)
        merged_text        # всё вместе — для анализа
    }
    """
    path = file_info["path"]
    ext  = "." + file_info["ext"].lower()

    result = {
        "path":              path,
        "ext":               file_info["ext"],
        "hash":              file_info["hash"],
        "raw_text":          "",
        "ocr_text":          "",
        "image_description": "",
        "merged_text":       "",
    }

    if ext in TEXT_EXTENSIONS:
        result["raw_text"] = _read_text(path)

    elif ext in SPREADSHEET_EXTENSIONS:
        result["raw_text"] = _extract_spreadsheet(path)

    elif ext in DOC_EXTENSIONS:
        if ext == ".pdf":
            result["raw_text"] = _extract_pdf_text(path)
        else:
            result["raw_text"] = _extract_with_docling(path)

        # Embedded изображения из PDF (другие форматы — в будущем)
        if ext == ".pdf" and config.get("process_embedded_images", True):
            embedded = _extract_pdf_images(path, config)
            if embedded:
                result["raw_text"] += "\n\n" + embedded

    elif ext in IMAGE_EXTENSIONS:
        if not config.get("process_standalone_images", True):
            return result  # только метаданные
        ocr, description = _call_qwen(path, config)
        result["ocr_text"]          = ocr
        result["image_description"] = description

    # Собрать merged_text
    parts = []
    if result["raw_text"]:
        parts.append(result["raw_text"])
    if result["ocr_text"]:
        parts.append(f"[OCR TEXT]\n{result['ocr_text']}")
    if result["image_description"]:
        parts.append(f"[IMAGE DESCRIPTION]\n{result['image_description']}")
    result["merged_text"] = "\n\n".join(parts)

    return result


# ---------------------------------------------------------------------------
# Текстовые файлы
# ---------------------------------------------------------------------------

def _read_text(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


# ---------------------------------------------------------------------------
# Документы через Docling
# ---------------------------------------------------------------------------

def _extract_with_docling(path: str) -> str:
    try:
        from docling.document_converter import DocumentConverter
    except ImportError:
        raise RuntimeError("Docling не установлен. Выполните: pip install docling")

    converter = DocumentConverter()
    result = converter.convert(path)
    return result.document.export_to_markdown()


def _extract_pdf_text(path: str) -> str:
    """Передаёт в Docling только первые _MAX_PAGES_PER_PDF страниц PDF."""
    import tempfile, os
    try:
        import fitz
    except ImportError:
        return _extract_with_docling(path)

    doc = fitz.open(path)
    page_count = len(doc)

    if page_count <= _MAX_PAGES_PER_PDF:
        doc.close()
        return _extract_with_docling(path)

    doc.select(list(range(_MAX_PAGES_PER_PDF)))
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp_path = tmp.name
        doc.save(tmp_path)
    doc.close()

    try:
        text = _extract_with_docling(tmp_path)
    finally:
        os.unlink(tmp_path)

    return text + f"\n\n[Показаны первые {_MAX_PAGES_PER_PDF} из {page_count} страниц]"


# ---------------------------------------------------------------------------
# Таблицы: .xlsx → openpyxl, .xls → xlrd
# ---------------------------------------------------------------------------

_MAX_ROWS_PER_SHEET = 500  # достаточно для понимания содержимого любой таблицы


def _extract_spreadsheet(path: str) -> str:
    ext = Path(path).suffix.lower()
    if ext == ".xls":
        return _extract_xls(path)
    return _extract_xlsx(path)


def _extract_xlsx(path: str) -> str:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl не установлен. Выполните: pip install openpyxl")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return _extract_with_docling(path)  # повреждённый файл
    parts = []
    for sheet in wb.worksheets:
        rows_text = _rows_to_text(sheet.iter_rows(values_only=True), sheet.max_row)
        if rows_text:
            parts.append(f"## {sheet.title}\n{rows_text}")
    wb.close()
    return "\n\n".join(parts)


def _extract_xls(path: str) -> str:
    try:
        import xlrd
    except ImportError:
        raise RuntimeError("xlrd не установлен. Выполните: pip install xlrd")
    try:
        wb = xlrd.open_workbook(path)
    except Exception:
        return _extract_with_docling(path)  # повреждённый файл
    parts = []
    for i in range(wb.nsheets):
        sheet = wb.sheet_by_index(i)
        rows_iter = (
            [sheet.cell(rx, cx).value for cx in range(sheet.ncols)]
            for rx in range(sheet.nrows)
        )
        rows_text = _rows_to_text(rows_iter, sheet.nrows)
        if rows_text:
            parts.append(f"## {sheet.name}\n{rows_text}")
    return "\n\n".join(parts)


def _rows_to_text(rows_iter, total_rows: int) -> str:
    result = []
    counted = 0
    for row in rows_iter:
        cells = [str(c).strip() for c in row if c is not None and str(c).strip() not in ("", "None")]
        if not cells:
            continue
        result.append("\t".join(cells))
        counted += 1
        if counted >= _MAX_ROWS_PER_SHEET:
            result.append(f"[... показаны первые {_MAX_ROWS_PER_SHEET} строк из {total_rows} ...]")
            break
    return "\n".join(result)


# ---------------------------------------------------------------------------
# Embedded-изображения из PDF через PyMuPDF
# ---------------------------------------------------------------------------

def _extract_pdf_images(path: str, config: dict) -> str:
    """
    Извлекает embedded-изображения из PDF, прогоняет каждое через Qwen3-VL.
    Возвращает строку с блоками описаний для вставки в raw_text.
    Изображения меньше _MIN_IMAGE_PX × _MIN_IMAGE_PX пропускаются.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise RuntimeError("PyMuPDF не установлен. Выполните: pip install PyMuPDF")

    doc = fitz.open(path)
    blocks = []
    img_count = 0
    total_imgs = sum(len(page.get_images(full=False)) for page in doc)

    for page_num, page in enumerate(doc, start=1):
        if page_num > _MAX_PAGES_PER_PDF:
            break
        if img_count >= _MAX_IMAGES_PER_PDF:
            break
        for img_info in page.get_images(full=False):
            if img_count >= _MAX_IMAGES_PER_PDF:
                blocks.append(f"[Показаны первые {_MAX_IMAGES_PER_PDF} из {total_imgs} изображений]")
                break

            xref = img_info[0]

            try:
                pix = fitz.Pixmap(doc, xref)

                # Пропустить слишком маленькие изображения
                if pix.width < _MIN_IMAGE_PX or pix.height < _MIN_IMAGE_PX:
                    continue

                # Конвертировать CMYK → RGB (Qwen3-VL ожидает RGB/RGBA)
                if pix.n - pix.alpha >= 4:
                    pix = fitz.Pixmap(fitz.csRGB, pix)

                img_bytes = pix.tobytes("png")

            except Exception:
                continue

            ocr, description = _call_qwen_bytes(img_bytes, config)

            if not ocr and not description:
                continue

            block_lines = [f"[Изображение, стр. {page_num}]"]
            if ocr:
                block_lines.append(f"OCR: {ocr}")
            if description:
                block_lines.append(f"Описание: {description}")
            blocks.append("\n".join(block_lines))
            img_count += 1

    doc.close()
    return "\n\n".join(blocks)


# ---------------------------------------------------------------------------
# Qwen3-VL через Ollama
# ---------------------------------------------------------------------------

def _call_qwen(path: str, config: dict) -> tuple:
    """Standalone-файл: читает с диска и отправляет в модель."""
    with open(path, "rb") as f:
        image_bytes = f.read()
    return _call_qwen_bytes(image_bytes, config)


def _resize_for_vlm(image_bytes: bytes, max_size: int) -> bytes:
    """Уменьшает изображение до max_size px по длинной стороне. При ошибке возвращает оригинал."""
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(image_bytes))
        w, h = img.size
        if max(w, h) <= max_size:
            return image_bytes
        scale = max_size / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        return buf.getvalue()
    except Exception:
        return image_bytes


def _call_qwen_bytes(image_bytes: bytes, config: dict) -> tuple:
    """Отправляет изображение (bytes) в Qwen3-VL. Возвращает (ocr_text, image_description)."""
    import requests

    max_size = config.get("image_max_size", _IMAGE_MAX_SIZE)
    image_bytes = _resize_for_vlm(image_bytes, max_size)

    prompt_path = Path(__file__).parent / "prompts" / "describe_image.txt"
    with open(prompt_path, "r", encoding="utf-8") as f:
        prompt = f.read()

    image_b64 = base64.b64encode(image_bytes).decode("utf-8")

    ollama = config["ollama"]
    response = requests.post(
        f"{ollama.get('base_url', 'http://localhost:11434')}/api/generate",
        json={
            "model":  ollama["model_name"],
            "prompt": prompt,
            "images": [image_b64],
            "stream": False,
        },
        timeout=ollama.get("timeout", 300),
    )
    response.raise_for_status()

    raw = response.json().get("response", "")
    return _parse_image_response(raw)


def _parse_image_response(raw: str) -> tuple:
    """Разобрать JSON из ответа модели → (ocr_text, image_description)."""
    try:
        data = json.loads(raw.strip())
        return str(data.get("ocr_text", "")), str(data.get("image_description", ""))
    except (json.JSONDecodeError, AttributeError):
        pass

    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if match:
        try:
            data = json.loads(match.group())
            return str(data.get("ocr_text", "")), str(data.get("image_description", ""))
        except json.JSONDecodeError:
            pass

    return "", raw.strip()
